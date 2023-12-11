import openpyxl

book = openpyxl.load_workbook("C:\\Users\\chitt\\PycharmProjects\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")

sheet = book.active  # to get the control to the active sheet
# in Excel row & column starts at 1.
cell = sheet.cell(row=1, column=2)  # Get the data of 1st row & 2nd column i.e. firstname
print(cell.value)  # Reads the cell value
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

Dict = {}

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
