import openpyxl


class HomePageData:
    test_HomePage_Data = [{"firstName": "Rakesh Sharma", "EmailID": "test1@gmail.com",
                           "Password": "test123", "Gender": "Male"},
                          {"firstName": "Kalpana Chawla", "EmailID": "test2@gmail.com",
                           "Password": "test123", "Gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):  # test_case_name it will come from the test case
        book = openpyxl.load_workbook(
            "C:\\Users\\chitt\\PycharmProjects\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")
        sheet = book.active  # to get the control to the active sheet
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # Compare with the TC name
                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            return[Dict]
